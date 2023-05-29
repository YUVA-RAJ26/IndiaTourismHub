from django.contrib.auth import authenticate, login, logout
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.authtoken.models import Token
from user_module.models import Users
from admin_module.models import State, TourPackage
from django.shortcuts import render, redirect
from django.shortcuts import get_object_or_404
from rest_framework.authtoken.models import Token
from .serializers import BookingSerializer
from rest_framework import viewsets, status
import io
from django.core.mail import EmailMessage
from django.template.loader import render_to_string
from django.http import FileResponse
from django.http import HttpResponse
from reportlab.pdfgen import canvas
from django.conf import settings
from .models import Booking
from rest_framework.decorators import action
from rest_framework.authentication import TokenAuthentication
from rest_framework.permissions import IsAuthenticated
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from io import BytesIO
from django.http import HttpResponse
from django.shortcuts import render
from django.contrib.admin.views.decorators import staff_member_required
import random
from user_module.tasks import send_otp_email_task
from django.core import serializers
import json
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.utils.html import strip_tags
import boto3

#IMPORT FROM UTILS
from user_module.utils.email import send_invoice_email, generate_invoice_pdf, send_booking_approved_email

def send_otp_email(request):
    email = request.POST.get('email')
    otp = random.randint(1000, 9999)
    user = Users.objects.get(email=email)
    user.otp = otp
    user.save()

    send_otp_email_task.delay(email, otp)
    return HttpResponse('success')

def login_page(request):
    return render(request, 'user_module/login.html')

def logout_page(request):
    logout(request)
    return redirect('login_page')

def sign_up(request):
    return render(request, 'user_module/signup.html')

def home(request):
    
    bookings = Booking.objects.all()
    approved_bookings = bookings.filter(payment_approved=True).count()
    rejected_bookings = bookings.filter(payment_rejected=True).count()
    pending_bookings = bookings.filter(payment_approved=False, payment_rejected=False).count()
    total_bookings = approved_bookings + rejected_bookings + pending_bookings

    approved_percentage = 0
    rejected_percentage = 0
    pending_percentage = 0
    if total_bookings != 0:
        approved_percentage = approved_bookings / total_bookings * 100
        rejected_percentage = rejected_bookings / total_bookings * 100
        pending_percentage = pending_bookings / total_bookings * 100

    # Data for the pie chart
    pie_data = [approved_percentage, rejected_percentage, pending_percentage]

    # Data for the bar chart
    bar_labels = ["Approved", "Rejected", "Pending"]
    bar_data = [approved_bookings, rejected_bookings, pending_bookings]

    states = State.objects.all()
    tourpackages = TourPackage.objects.all()
    token = Token.objects.get(user=request.user)
    context = {
        'authToken': f'Token {str(token)}',
        'pie_labels': bar_labels,
        'pie_data': pie_data,
        'bar_labels': bar_labels,
        'bar_data': bar_data,
        'states': states,
        'tourpackages': tourpackages,
    }

    return render(request, 'user_module/home.html', context)

def booking_payment_update(request):
    token = Token.objects.get(user=request.user)
    pk = request.GET.get('pk')
    booking = Booking.objects.get(pk=pk)
    context = {
        'booking': booking,
        'authToken': f'Token {str(token)}',
    }
    return render(request, 'booking/booking_payment_update.html', context)

def send_invoice_email(booking, invoice_pdf):
    # Render email template with booking details
    email_subject = 'Invoice for Booking'
    email_body = 'Hi! Happy travelling Please find the invoice! Make a payment and update to confrom booking'
    to = [booking.user.email]

    # Create the email message
    email_message = EmailMessage(
        email_subject,
        email_body,
        to=to,
    )
    
    # Attach the invoice PDF to the email
    email_message.attach('invoice.pdf', invoice_pdf.getvalue(), 'application/pdf')

    #email_message.send()

    # Send the email
    try:
        email_message.send()
    except Exception as e:
        return True
    return True

def generate_invoice_pdf(booking):
    buffer = io.BytesIO()

    # Create the PDF object
    p = canvas.Canvas(buffer)

    # Customize and draw the invoice contents
    p.setFont("Helvetica", 12)
    p.drawString(50, 750, "Invoice for Booking ID: {}".format(booking.id))
    p.drawString(50, 700, "Customer Name: {}".format(booking.user.username))
    p.drawString(50, 650, "Tour Package: {}".format(booking.tour_package.name))
    p.drawString(50, 600, "Booking Date: {}".format(booking.book_date.strftime("%Y-%m-%d")))
    p.drawString(50, 550, "Contact Number: {}".format(booking.contact_number))
    p.drawString(50, 450, "Number of Adults: {}".format(booking.num_adults))
    p.drawString(50, 400, "Number of Children: {}".format(booking.num_children))
    p.drawString(50, 350, "Price: ${}".format(booking.price))

    # Save the PDF
    p.showPage()
    p.save()

    # Set the buffer position to the beginning
    buffer.seek(0)

    return buffer

def download_charts_as_pdf(request):
    # Get the bookings data and calculate the necessary values for the charts
    bookings = Booking.objects.all()
    approved_bookings = bookings.filter(payment_approved=True).count()
    rejected_bookings = bookings.filter(payment_rejected=True).count()
    pending_bookings = bookings.count() - approved_bookings - rejected_bookings

    # Define the data for the pie chart
    pie_labels = ['Approved', 'Rejected', 'Pending']
    pie_data = [approved_bookings, rejected_bookings, pending_bookings]

    # Define the data for the bar chart
    bar_labels = ['Approved', 'Rejected', 'Pending']
    bar_data = [approved_bookings, rejected_bookings, pending_bookings]

    # Create the pie chart
    fig1, ax1 = plt.subplots()
    ax1.pie(pie_data, labels=pie_labels, autopct='%1.1f%%', startangle=90)
    ax1.axis('equal')

    # Create the bar chart
    fig2, ax2 = plt.subplots()
    ax2.bar(bar_labels, bar_data)

    # Create a PDF document and add the charts
    pdf_buffer = BytesIO()
    with PdfPages(pdf_buffer) as pdf:
        pdf.savefig(fig1)
        pdf.savefig(fig2)

    # Set the buffer's position to the beginning
    pdf_buffer.seek(0)

    # Generate the response with the PDF file
    response = HttpResponse(pdf_buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="booking_charts.pdf"'

    return response

def send_booking_approved_email(booking):
    subject = 'Booking Approved'
    html_message = render_to_string('booking/booking_approved.html', {'booking': booking})
    plain_message = strip_tags(html_message)
    from_email = settings.DEFAULT_FROM_EMAIL
    to_email = booking.user.email

    send_mail(subject, plain_message, from_email, [to_email], html_message=html_message)


import openpyxl
from django.http import HttpResponse

def download_bookings_excel(request):
    # Get the filter parameters from the request (e.g., package ID, payment status)
    package_id = request.GET.get('package_id')
    payment_status = request.GET.get('payment_status')

    # Filter the bookings based on the selected parameters
    bookings = Booking.objects.all()
    if package_id:
        bookings = bookings.filter(package_id=package_id)
    if payment_status:
        if payment_status == 'false':
            bookings = bookings.filter(payment_approved=False)
        elif payment_status == 'true':
            bookings = bookings.filter(payment_approved=True)

    # Create a new Excel workbook and worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Write the headers
    headers = ['Booking ID', 'User', 'Tour Package', 'Booking Date', 'Contact Number', 'Payment Approved']
    for col_num, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col_num, value=header)

    # Write the booking details
    for row_num, booking in enumerate(bookings, 2):
        worksheet.cell(row=row_num, column=1, value=booking.id)
        worksheet.cell(row=row_num, column=2, value=booking.user.username)
        worksheet.cell(row=row_num, column=3, value=booking.tour_package.name)
        worksheet.cell(row=row_num, column=4, value=booking.book_date)
        worksheet.cell(row=row_num, column=5, value=booking.contact_number)
        worksheet.cell(row=row_num, column=6, value='Completed' if booking.payment_approved else 'Pending')

    # Create a response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=bookings.xlsx'
    workbook.save(response)

    return response

# Listing HTML Pages Calls
def booking_list(request):
    token = Token.objects.get(user=request.user)
    context = {
        'authToken': f'Token {str(token)}',
        'packages': TourPackage.objects.all(),
    }
    return render(request, 'booking/booking_list.html', context)

#Start Of Class Based Views
class LoginView(APIView):

    def get(self, request):
        pass
    
    def post(self, request):
        otp = request.data.get('otp')
        email = request.data.get('email')
        password = request.data.get('password')
        username = request.data.get('username')

        # Authenticate user
        user = authenticate(request, email=email, password=password, otp=otp)
        
        if user is not None:
            # Log the user in
            login(request, user)
            token, _ = Token.objects.get_or_create(user=user)
            return Response({'message': 'Logged in successfully!'}, status=status.HTTP_200_OK)
        else:
            return Response({'message': 'Invalid email or password.'}, status=status.HTTP_400_BAD_REQUEST)
        
class SignUpView(APIView):
    def post(self, request):
        username = request.data.get('username')
        password = request.data.get('password')
        email = request.data.get('email')
        contact_no = request.data.get('contact_no')

        try:
            user = Users.objects.get(email=email) 
            return Response({'message': 'Email already exists.'}, status=status.HTTP_400_BAD_REQUEST)
        except Users.DoesNotExist:            
            user = Users.objects.create_user(
                username=username,
                password=password,
                email=email,
                contact_no=contact_no,
            )
            
            return Response({'message': 'Signup successful'}, status=status.HTTP_201_CREATED)

class BookingViewSet(viewsets.ModelViewSet):
    serializer_class = BookingSerializer
    queryset = Booking.objects.all()  # Add the queryset attribute
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def list(self, request, *args, **kwargs):
        payment_status = request.GET.get('payment_status', '')
        package = request.GET.get('package', '')

        if request.user.is_staff:
            queryset = Booking.objects.all()
        else:
            queryset = Booking.objects.filter(user=request.user)

        # Apply filters
        if payment_status:
            queryset = queryset.filter(payment_approved=(payment_status.lower() == 'true'))
        if package:
            queryset = queryset.filter(tour_package=package)  

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)
    
    def create(self, request):
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            booking = serializer.save()

            # Generate invoice PDF
            invoice_pdf = generate_invoice_pdf(booking)

            # Send invoice via email
            send_invoice_email(booking, invoice_pdf)

            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, pk=None):
        instance = Booking.objects.get(pk=pk)

        if 'payment_screenshot' in request.data:
            instance.payment_screenshot = request.data['payment_screenshot']

        instance.save()

        # Serialize the updated fields and return the response
        serializer = self.get_serializer(instance)
        response_data = {
            'data': serializer.data
        }
        self.send_admin_notification(instance)
        return Response(response_data)            

    def send_admin_notification(self, booking):
        # Create the email message
        subject = "New Payment Screenshot Uploaded"
        body = "A new payment screenshot has been uploaded for Booking ID: {}".format(booking.id)
        from_email = settings.DEFAULT_FROM_EMAIL
        to_email = settings.EMAIL_HOST_USER  # Replace with the admin's email address

        email = EmailMessage(subject, body, from_email, [to_email])

        # Attach the payment screenshot file to the email
        if booking.payment_screenshot:
            email.attach_file(booking.payment_screenshot.path)

        email.send()
    
    @action(detail=True, methods=['put'])
    def mark_completed(self, request, pk=None):
        booking = Booking.objects.get(id=pk)
        print(booking.payment_approved)
        approved = request.data.get('payment_approved')
        if approved == 'true':
            booking.payment_approved = True 
            booking.payment_rejected = False
            booking.save()
            # Send email notification
            send_booking_approved_email(booking)

        else:
            booking.payment_approved = False
            booking.payment_rejected = True

        booking.save()
        print(booking.payment_approved)
        serializer = self.get_serializer(booking)
        return Response(serializer.data)

    @action(detail=False, methods=["GET"], url_path="get_state_districts", url_name="get_state_districts")
    def get_context_data(self, request):
        selected_state = request.GET.get('state')
        selected_tour_package = request.GET.get('tourPackage')

        bookings = Booking.objects.all()

        # Apply filters based on selected state and tour package
        if selected_state != 'all':
            bookings = bookings.filter(tour_package__tourist_place__state__id=selected_state)
        if selected_tour_package != 'all':
            bookings = bookings.filter(tour_package__id=selected_tour_package)
        
        print(bookings)

        approved_bookings = bookings.filter(payment_approved=True).count()
        rejected_bookings = bookings.filter(payment_rejected=True).count()
        pending_bookings = bookings.filter(payment_approved=False, payment_rejected=False).count()
        total_bookings = approved_bookings + rejected_bookings + pending_bookings


        approved_percentage = 0
        rejected_percentage = 0
        pending_percentage = 0
        if total_bookings != 0:
            approved_percentage = approved_bookings / total_bookings * 100
            rejected_percentage = rejected_bookings / total_bookings * 100
            pending_percentage = pending_bookings / total_bookings * 100

        # Data for the pie chart
        pie_data = [approved_percentage, rejected_percentage, pending_percentage]

        # Data for the bar chart
        bar_labels = ["Approved", "Rejected", "Pending"]
        bar_data = [approved_bookings, rejected_bookings, pending_bookings]

        states = serializers.serialize('json', State.objects.all())
        tourpackages = serializers.serialize('json', TourPackage.objects.all())

        context = {
            'pie_labels': bar_labels,
            'pie_data': pie_data,
            'bar_labels': bar_labels,
            'bar_data': bar_data,
            'states': states,
            'tourpackages': tourpackages,
        }

        return Response(context)




#PAYMENT IMPLEMENTATION
# views.py
from django.shortcuts import render
from django.conf import settings
from paytmchecksum import PaytmChecksum

# Counter for generating sequential IDs
order_id_counter = 100000000000000000
customer_id_counter = 100000000000000000

def generate_order_id():
    global order_id_counter
    order_id_counter += 1
    return f'order_{order_id_counter}'

def generate_customer_id():
    global customer_id_counter
    customer_id_counter += 1
    return f'customer_{customer_id_counter}'

def initiate_payment(request):
    # Get payment parameters from the request or generate them dynamically
    params = {
        'MID': settings.PAYTM_MERCHANT_ID,
        'WEBSITE': settings.PAYTM_WEBSITE,
        'INDUSTRY_TYPE_ID': 'Retail',
        'CHANNEL_ID': 'WEB',
        'ORDER_ID': generate_order_id(),
        'CUST_ID': generate_customer_id(),
        'TXN_AMOUNT': '10.00',  # Replace with the actual transaction amount
        'CALLBACK_URL': 'http://localhost:8000/user-module/payment/callback/',
        # Add any additional parameters as required by Paytm
    }

    # import pdb
    # pdb.set_trace()  
    # Generate a checksum for the payment parameters
    checksum = PaytmChecksum.generateSignature(params, settings.PAYTM_MERCHANT_KEY)

    # Pass the payment parameters and checksum to the template
    context = {
        'params': params,
        'checksum': checksum,
    }

    return render(request, 'payment.html', context)

# Rest of the code remains the same




def payment_callback(request):
    # Retrieve the payment response parameters from the request
    received_data = dict(request.POST)

    # Verify the checksum of the payment response
    is_checksum_valid = PaytmChecksum.verifySignature(received_data, settings.PAYTM_MERCHANT_KEY)

    if is_checksum_valid:
        # Process the payment response
        # Update order status, record transaction details, etc.
        # Access the payment response parameters using received_data dictionary
        # Example: payment_status = received_data.get('STATUS')

        return render(request, 'payment_success.html')
    else:
        return render(request, 'payment_failure.html')
# views.py
from django.shortcuts import render
from django.conf import settings
import stripe
import traceback
import stripe
from stripe.error import CardError
from stripe.util import convert_to_stripe_object
import stripe
from stripe.error import CardError

stripe.api_key = settings.STRIPE_SECRET_KEY

def proceed_payment(request):
    if request.method == 'POST':
        # Get the payment method details from the form
        payment_method = request.POST.get('payment_method')

        try:
            # Create or retrieve the PaymentIntent
            payment_intent = stripe.PaymentIntent.create(
                amount=1,  # Replace with the actual amount
                currency='usd',  # Replace with the actual currency
                payment_method=payment_method,
                confirmation_method='manual',
                confirm=True
            )

            # Pass the payment_intent object as context to the template
            return render(request, 'payment/proceed.html', {'payment_intent': payment_intent})
        except stripe.error.StripeError as e:
            # Handle any Stripe API errors
            return render(request, 'payment/error.html', {'error': str(e)})

    return render(request, 'payment/proceed.html')





# views.py

from django.http import HttpResponse
import stripe
import json

stripe.api_key = settings.STRIPE_SECRET_KEY

def webhook(request):
    payload = request.body
    event = None

    try:
        event = stripe.Webhook.construct_event(
            payload, request.META.get('HTTP_STRIPE_SIGNATURE'), settings.STRIPE_WEBHOOK_SECRET
        )
    except ValueError as e:
        # Invalid payload
        return HttpResponse(status=400)
    except stripe.error.SignatureVerificationError as e:
        # Invalid signature
        return HttpResponse(status=400)

    # Handle specific event types
    if event.type == 'payment_intent.succeeded':
        payment_intent = event.data.object
        # Handle payment success

    return HttpResponse(status=200)
